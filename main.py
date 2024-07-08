import pandas as pd
from icalendar import Calendar
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def convert_to_datetime(date_obj):
    if isinstance(date_obj, datetime.date) and not isinstance(date_obj, datetime.datetime):
        return datetime.datetime.combine(date_obj, datetime.datetime.min.time())
    return date_obj

def format_datetime(datetime_obj):
    if datetime_obj is not None:
        return datetime_obj.strftime('%d-%m-%Y %H:%M')
    return None

def format_rrule(rrule):
    if rrule is not None and 'FREQ' in rrule:
        freq = rrule['FREQ']
        if freq == ['DAILY'] and 'UNTIL' in rrule:
            until_date = rrule['UNTIL'][0]  # Access the first element of the list
            until_date_str = until_date.strftime('%d-%m-%Y %H:%M')
            return f"Tous les jours jusqu'au {until_date_str}"
        if freq == ['WEEKLY'] and 'UNTIL' in rrule:
            until_date = rrule['UNTIL'][0]  # Access the first element of the list
            until_date_str = until_date.strftime('%d-%m-%Y %H:%M')
            return f"Toutes les semaines jusqu'au {until_date_str}"
        if freq == ['WEEKLY'] and 'UNTIL' in rrule:
            until_date = rrule['UNTIL'][0]  # Access the first element of the list
            until_date_str = until_date.strftime('%d-%m-%Y %H:%M')
            return f"Tous les mois jusqu'au {until_date_str}"
    return None

def apply_alternating_row_colors(ws):
    # Apply alternating row colors to all rows (excluding header)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.row % 2 == 0:  # Even row
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:  # Odd row
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def adjust_column_widths(ws, width_in_pixels=300):
    # Adjust column widths for all columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

def ics_to_excel(calendar_files, excel_file):
    all_events = []

    # Prompt user for filtering options
    filter_summary_input = input("Do you want to filter by summary? (yes/no): ")
    filter_summary = filter_summary_input.lower() == 'yes'

    if filter_summary:
        keyword = input("Enter the keyword to filter by: ").strip()
    else:
        keyword = None

    for calendar_file in calendar_files:
        calendar_name = os.path.splitext(os.path.basename(calendar_file))[0]  # Extract calendar name from file
        with open(calendar_file, 'r', encoding='utf-8') as f:
            calendar = Calendar.from_ical(f.read())

        events = []
        for component in calendar.walk():
            if component.name == "VEVENT":
                dtstart = component.decoded('dtstart') if component.get('dtstart') else None
                dtend = component.decoded('dtend') if component.get('dtend') else None
                rrule = component.get('rrule') if component.get('rrule') else None
                summary = component.get('summary') if component.get('summary') else None

                # Convert to timezone-unaware if they are not None
                if dtstart is not None:
                    dtstart = convert_to_datetime(dtstart)
                    if hasattr(dtstart, 'tzinfo'):
                        dtstart = dtstart.replace(tzinfo=None)
                
                if dtend is not None:
                    dtend = convert_to_datetime(dtend)
                    if hasattr(dtend, 'tzinfo'):
                        dtend = dtend.replace(tzinfo=None)
                
                # Parse and format rrule if available
                rrule_formatted = format_rrule(rrule)

                # Determine the title to display based on the filter condition
                if filter_summary and keyword and summary:
                    if keyword.lower() not in summary.lower():
                        summary_to_display = None  # Do not display this event in output
                    else:
                        summary_to_display = summary
                else:
                    summary_to_display = summary

                event = {
                    'Formateur': calendar_name,
                    'Debut': dtstart,
                    'Fin': dtend,
                    'Récurrence': rrule_formatted,
                    'Titre': summary_to_display,
                }
                events.append(event)
        all_events.extend(events)

    # Create a DataFrame
    df = pd.DataFrame(all_events)

    # Filter out events that start before now
    now = datetime.datetime.now()
    df = df[df['Debut'] >= now]

    # Sort by 'dtstart'
    df = df.sort_values(by='Debut')

    # Format datetime fields
    df['Debut'] = df['Debut'].apply(format_datetime)
    df['Fin'] = df['Fin'].apply(format_datetime)

    # Export to Excel
    df.to_excel(excel_file, index=False)

    # Load workbook and apply adjustments
    wb = load_workbook(excel_file)
    ws = wb.active

    # Apply alternating row colors
    apply_alternating_row_colors(ws)

    # Adjust column widths
    adjust_column_widths(ws)

    # Save workbook
    wb.save(excel_file)

def main():
    # Define calendar files and output Excel file
    calendar_files = [
        'exemple1.ics',
        'exemple2.ics',
    ]

    excel_file = 'output_file.xlsx'

    # Call function to process ICS files
    ics_to_excel(calendar_files, excel_file)

if __name__ == "__main__":
    main()
